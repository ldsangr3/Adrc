# Release ADRC 
# Stils Bugs at vector of DO of PBR3
import numpy as np #Numpy
from ADRC import *
import tkinter as tk #TKINTER ES PARA INTERFAZ GRÁFICA
from tkinter import ttk # TTK



import matplotlib.pyplot as plt #GRAHP LIBRARY
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg #GUI CANVAS
from scipy.signal import butter, lfilter

# Image loading library
from PIL import ImageTk, Image

from ljm1 import Ljm1 #LabJack1
from ljm2 import Ljm2 #LabJack2

import continuous_threading
import time
import datetime

from PID import PID_Event_Based
from PID_Continuo import PID_Realtime
from PID_Discreto import PID_Discreto
from ADRC import ADRC
from ADRC_II import ADRC_II
from Models import Model_Microalgae


# Library to export data
import xlsxwriter
import openpyxl

# Library to append data avoiding diferences in sizes
from collections import deque


# Intances for the Labjacks, and starting the comunication
Labjack1 = Ljm1() #LIB
Labjack2 = Ljm2() #LIB

class Window_FBR:
    def __init__(self,tab):
        """
        New Widows 
        """
        right_frame = tk.Frame(master=tab, bg='#00205B', bd=1.5) #bg=color; bd=tamaño del borde en pixeles # Care master add
        right_frame.place(relx=0.3, rely=0.055, relwidth=0.65, relheight=0.8) #posicionamiento de elementos horizontal,vertical,ancho,largo
        notebook=ttk.Notebook(right_frame)
        notebook.pack(fill='both',expand='yes')
        tab_temperature=ttk.Frame(notebook)
        tab_Light=ttk.Frame(notebook)
        tab_PBR_lvl=ttk.Frame(notebook)
        tab_DO=ttk.Frame(notebook)
        tab_pH=ttk.Frame(notebook)
        tab_Biomass=ttk.Frame(notebook)
        notebook.add(tab_temperature,text=' Temperature ')
        notebook.add(tab_Light,text=' Light Intensity ')
        notebook.add(tab_PBR_lvl,text=' PBR Level ')       
        notebook.add(tab_DO,text=' DO ')
        notebook.add(tab_pH,text=' pH ')
        notebook.add(tab_Biomass,text=' Biomass ')
        
        # Mini panels for the tab
        tk.Label(tab,text="Light",font=("Helvetica",15), fg='white', bg=('#00205B')).place(x=10,y=10)
        tk.Label(tab,text="Color",font=("Helvetica",14), fg='white', bg=('#00205B')).place(x=10,y=40)
        
        #   Configuration for the panels for the varaibles 
        
        frame_variable_temp= plt.Figure(figsize=(5,6), dpi=100)
        frame_variable_Iin= plt.Figure(figsize=(5,6), dpi=100)
        frame_variable_Lvl= plt.Figure(figsize=(5,6), dpi=100)
        frame_variable_pH= plt.Figure(figsize=(5,6), dpi=100)
        frame_variable_DO= plt.Figure(figsize=(5,6), dpi=100)
        frame_variable_Biomass= plt.Figure(figsize=(5,6), dpi=100)
        
        
        self.xTemp = frame_variable_temp.add_subplot(111)
        self.xIin = frame_variable_Iin.add_subplot(111)
        self.xLvl = frame_variable_Lvl.add_subplot(111)
        self.xpH = frame_variable_pH.add_subplot(111)
        self.xDO = frame_variable_DO.add_subplot(111)
        self.xBiomas = frame_variable_Biomass.add_subplot(111)
        
        #Figure Temperature
        self.xTemp.grid(True),self.xTemp.set_xlabel('$Time$'),self.xTemp.set_ylabel('$°C$')
        self.xTemp.set_ylim([0, 50])
        self.lineTemp = FigureCanvasTkAgg(frame_variable_temp, tab_temperature)
        self.lineTemp.get_tk_widget().pack(side=tk.LEFT, fill=tk.BOTH,expand=1)
        
        #Figure Ligt
        self.xIin.grid(True),self.xIin.set_xlabel('$Time$'),self.xIin.set_ylabel('$\mu mol \cdot m^{-2} \cdot s^{-1}$')
        self.xIin.set_ylim([0, 600])
        self.lineIin = FigureCanvasTkAgg(frame_variable_Iin, tab_Light)
        self.lineIin.get_tk_widget().pack(side=tk.LEFT, fill=tk.BOTH,expand=1)
        
        # Figure level
        self.xLvl.grid(True),self.xLvl.set_xlabel('$Time$'),self.xLvl.set_ylabel('$cm$')
        
        self.lineLvl = FigureCanvasTkAgg(frame_variable_Lvl, tab_PBR_lvl)
        self.lineLvl.get_tk_widget().pack(side=tk.LEFT, fill=tk.BOTH,expand=1)
        
        #FIGURA pH
        self.xpH.grid(True),self.xpH.set_xlabel('$Time$'),self.xpH.set_ylabel('$pH$')
        self.xpH.set_ylim([0, 14])
        self.linepH = FigureCanvasTkAgg(frame_variable_pH, tab_pH)
        self.linepH.get_tk_widget().pack(side=tk.LEFT, fill=tk.BOTH,expand=1)
        
        #FIGURA DO
        self.xDO.grid(True),self.xDO.set_xlabel('$TIme$'),self.xDO.set_ylabel('$mg \cdot L^{-1}$')
        self.xDO.set_ylim([0, 100])
        self.lineDO = FigureCanvasTkAgg(frame_variable_DO, tab_DO)
        self.lineDO.get_tk_widget().pack(side=tk.LEFT, fill=tk.BOTH,expand=1)
        
        #FIGURA Biomass
        self.xBiomas.grid(True),self.xBiomas.set_xlabel('$Time$'),self.xBiomas.set_ylabel('$Density$')
        self.xBiomas.set_ylim([0, 100])
        self.lineBiomass = FigureCanvasTkAgg(frame_variable_Biomass, tab_Biomass)
        self.lineBiomass.get_tk_widget().pack(side=tk.LEFT, fill=tk.BOTH,expand=1)
        
        # Initial values of ports 
        #These pings are conected to Int port of the PMPs 
        Labjack2.sendValue('TDAC2', 0)  
        Labjack2.sendValue('TDAC3', 0)
        Labjack2.sendValue('TDAC0', 0)
        Labjack2.sendValue('TDAC1', 0)
        Labjack2.sendValue('DAC0', 0)
        Labjack2.sendValue('DAC1', 0)
        
    
    def __main__(self, tab):
        
        self.tab = tab #intance of tab
        
        self.rcolor = tk.IntVar(master=tab,value=4)
        tab.rcolor = tk.Radiobutton(tab, text="Red", variable=self.rcolor, value=1, command=self.color_rojo,font=("Helvetica",10), bg=('#00205B'), fg=('white'), activebackground='#0E62A6', activeforeground='white', selectcolor='black', foreground='white').place(x=10,y=70)
        tab.rcolor = tk.Radiobutton(tab, text="Green", variable=self.rcolor, value=2, command=self.color_verde,font=("Helvetica",10), bg=('#00205B'), fg=('white'), activebackground='#0E62A6', activeforeground='white', selectcolor='black', foreground='white').place(x=10,y=95)
        tab.rcolor = tk.Radiobutton(tab, text="Blue", variable=self.rcolor, value=3, command=self.color_azul,font=("Helvetica",10), bg=('#00205B'), fg=('white'), activebackground='#0E62A6', activeforeground='white', selectcolor='black', foreground='white').place(x=10,y=120)
        tab.rcolor = tk.Radiobutton(tab, text="White", variable=self.rcolor, value=4, command=self.color_white,font=("Helvetica",10), bg=('#00205B'), fg=('black'), activebackground='#0E62A6', activeforeground='white', selectcolor='black', foreground='white').place(x=10,y=145)
        
        #activebackground=common_bg, activeforeground=common_fg, selectcolor=common_bg
               
                            
        self.Ref_luz=tk.DoubleVar(master=tab)
        tk.Scale(tab,variable=self.Ref_luz, from_ = 1, to = 900, orient = "horizontal",length=217,bg=('#BDBDBD')).place(x=80,y=70)
        
        ttk.Entry(tab, width=31, textvariable=self.Ref_luz,font=("Helvetica",10)).place(x=80,y=110)
        tk.Label(tab,text="Color Intensity",font=("Helvetica",15), fg='white', bg=('#00205B')).place(x=140,y=40)
        tk.Label(tab,text="Temperature",font=("Helvetica",15), fg='white', bg=('#00205B')).place(x=10,y=175)
        tk.Label(tab, text="Ref:",font=("Helvetica",12), fg='white', bg=('#00205B')).place(x=10,y=205)
        
        self.Ref_tmp=tk.DoubleVar(master=tab)
        tk.Scale(tab,variable=self.Ref_tmp, from_ = 1, to = 92, orient = "horizontal",length=217,bg=('#BDBDBD')).place(x=80,y=205)
        ttk.Entry(tab, width=31, textvariable=self.Ref_tmp, font=("Helvetica",10)).place(x=80,y=245)
        tk.Label(tab,text="PBR Level",font=("Helvetica",15), fg='white',bg=('#00205B')).place(x=10,y=275)
        tk.Label(tab, text="Ref:",font=("Helvetica",12),fg='white', bg=('#00205B')).place(x=10,y=305)
        
        self.Ref_lvl=tk.DoubleVar(master=tab)
        tk.Scale(tab,variable=self.Ref_lvl, from_ = 1, to = 25, orient = "horizontal",length=217,bg=('#BDBDBD')).place(x=80,y=305)
        ttk.Entry(tab, width=31, textvariable=self.Ref_lvl,font=("Helvetica",10)).place(x=80,y=345)
        
        

   

            
    # Color activation    
    def color_rojo(self): 
        return
    def color_verde(self):   
        return
    def color_azul(self):
        return
    def color_white(self):
        #PBR1
        Labjack1.sendValue('EIO0', True)  # 1
        Labjack1.sendValue('EIO1', False) # 2
        Labjack1.sendValue('EIO2', False) # 3   
        #PBR2
        Labjack1.sendValue('EIO3', False) # 1
        Labjack1.sendValue('EIO4', True)  # 2
        Labjack1.sendValue('EIO5', False) # 3 
        #PBR3
        Labjack1.sendValue('EIO6', True)  # 1
        Labjack1.sendValue('EIO7', False) # 2
        Labjack1.sendValue('CIO0', False) # 3
        


class SettingWidnows:
    def __init__(self):
        a = 1
        #hola

class threads:
    pass

class Main():

# Main funtions
    
    def __init__(self):
        self.root = tk.Tk()
        
        # Icono
        self.root.tk.call('wm', 'iconphoto', self.root._w, tk.PhotoImage(file='supervision.png'))
        # Configuratin of the wimdows
        self.root.title("ADRC")
        self.root.geometry('1280x960') #Size pixels
        
        
        menubar = tk.Menu(self.root) # New Menu
        self.root.config(menu=menubar)  
        
        # Menu for File
        menu_archivo = tk.Menu(menubar) #ELEMENTOS BARRA DE MENU
        menu_archivo = tk.Menu(menubar, tearoff=0)    
        menu_edicion = tk.Menu(menubar)
        menu_edicion = tk.Menu(menubar, tearoff=0)
        menu_ayuda = tk.Menu(menubar)
        menu_ayuda = tk.Menu(menubar, tearoff=0)
        
        # Menus
        menubar.add_cascade(label="File", menu=menu_archivo)
        menubar.add_cascade(label="Edit", menu=menu_edicion)
        menubar.add_cascade(label="Help", menu=menu_ayuda)
                
        # Comands for menus
        menu_archivo.add_command(label="New")
        menu_archivo.add_command(label="Open")
        menu_archivo.add_command(label="Save")
        menu_archivo.add_separator()
        menu_archivo.add_command(label="Exit", command=self.close)
        
        menu_edicion.add_command(label="Reset PID PBR1", command=self.PID_tmp_PBR1reset)
        menu_edicion.add_command(label="Reset PID PBR2", command=self.PID_tmp_PBR2reset)
        menu_edicion.add_command(label="Reset PID PBR3", command=self.PID_tmp_PBR3reset)
        
        
        menu_ayuda.add_command(label="Thanks", command=self.f_acerca) # f.acerca = MÉTODO CON LA FIGURA AGRADECIMIENTO
        menu_ayuda.add_command(label="About", command=self.f_about)   # f.about  = About method
        
        #
        self.notebook=ttk.Notebook(self.root) #CREAR OBJETO PARA GESTIONAR LAS PESTAÑAS
        self.notebook.pack(fill='both',expand='yes') # COnfiguration para que las pestañas tengan el mismo tamaño de la pant
        
              
        # Create a photoimage object of the image in the path
        icono_sabana = ImageTk.PhotoImage(Image.open("MONOCROMATICA-HRZL.png").resize((176, 67)))
        
        
        
        # Create a label object to display the image   
        label_Sabana = ttk.Label(image= icono_sabana)
        label_Sabana.image =  icono_sabana
        # Position imagex
        label_Sabana.place(x=700, y=790)

        # Blue Frames
        self.tab_FBR1=tk.Frame(self.notebook,bg='#00205B')  #
        self.tab_FBR2=tk.Frame(self.notebook,bg='#00205B')  #
        self.tab_FBR3=tk.Frame(self.notebook,bg='#00205B')  #
        
        
        
        # Tabs FBR1,FBR2 Y FBR3
        self.notebook.add(self.tab_FBR1,text='FBR1') #
        self.notebook.add(self.tab_FBR2,text='FBR2') #
        self.notebook.add(self.tab_FBR3,text='FBR3') #     
        
        # Configuration of FBR1, FBR2 and FBR3
        self.wFBR1 = Window_FBR(self.tab_FBR1)     
        self.wFBR2 = Window_FBR(self.tab_FBR2)  
        self.wFBR3 = Window_FBR(self.tab_FBR3)
        
        # Threads Execution times
        self.pid_temperature_executiontime=2 #Time of execution fo the temperature thread
        self.Nivel_executiontime=0.1 #Time of execution fo the Nivel thread
        self.Light_executiontime=1 #Time of execution fo the temperature thread
        self.Light_Control_executiontime=120 #Time of execution fo the temperature thread
        
        # Instances ofs PIDs for temperature
        #self.PID_temp_PBR1 = PID_Event_Based (P=10*6.48, I=20, D=2, Z=self.pid_temperature_executiontime)
        self.PID_temp_PBR1 = PID_Event_Based (P=0.0719593634132244, I=0.0205325079870776, D=-0.584407891678162, Z=self.pid_temperature_executiontime)
        self.PID_temp_PBR2 = PID_Event_Based (P=0.0719593634132244, I=0.0205325079870776, D=-0.584407891678162, Z=self.pid_temperature_executiontime)
        self.PID_temp_PBR3 = PID_Event_Based (P=0.0719593634132244, I=0.0205325079870776, D=-0.584407891678162, Z=self.pid_temperature_executiontime)  
        
        # # Instances ofs PIDs for Ligth
        self.PID_Light_PBR1 = PID_Discreto ()
        self.PID_Light_PBR2 = PID_Discreto ()
        self.PID_Light_PBR3 = PID_Discreto ()
        
        
        # Start the Initial values of the Light
        
        Labjack1.sendValue('TDAC3', 3.5)
        Labjack1.sendValue('TDAC4', 3.5)
        Labjack1.sendValue('TDAC5', 3.5)
        
        # Start 
        self.avfiltered_Iin_PBR1= [1] * 10
        self.avfiltered_Iin_PBR2= [1] * 10
        self.avfiltered_Iin_PBR3= [1] * 10
        
        
        #Start storage variables
        # Global variables 
        _MAX_size_ = 50  # Maximum size of the list
        self.Time_PBR1_pH=deque([datetime.datetime.now()], maxlen=_MAX_size_)
        self.Time_PBR1_DO=deque([datetime.datetime.now()], maxlen=_MAX_size_)
        self.Time_PBR2_pH=deque([datetime.datetime.now()], maxlen=_MAX_size_)
        self.Time_PBR2_DO=deque([datetime.datetime.now()], maxlen=_MAX_size_)
        self.Time_PBR3_pH=deque([datetime.datetime.now()], maxlen=_MAX_size_)
        self.Time_PBR3_DO=deque([datetime.datetime.now()], maxlen=_MAX_size_)
        
                       
        self.pH_PBR1=deque([0], maxlen=_MAX_size_)
        self.pH_PBR2=deque([0], maxlen=_MAX_size_)
        self.pH_PBR3=deque([0], maxlen=_MAX_size_)
        
        self.DO_PBR1=deque([0], maxlen=_MAX_size_)
        self.DO_PBR2=deque([0], maxlen=_MAX_size_)
        self.DO_PBR3=deque([0], maxlen=_MAX_size_)
        
        self.lvl_PBR1=[]
        self.lvl_PBR2=[]
        self.lvl_PBR3=[]
        
        self.Timelvl_PBR1=[]
        self.Timelvl_PBR2=[]
        self.Timelvl_PBR3=[]
        
        self.Iin_PBR1=[]
        self.Iin_PBR2=[]
        self.Iin_PBR3=[]
        
        self.TimeIin=[]
        self.average_Iin_PBR1=[]
        self.average_Iin_PBR2=[]
        self.average_Iin_PBR3=[]
        
        self.TimeTemperature=[]
        
        self.Tmp_PBR1=[]
        self.Tmp_PBR2=[]
        self.Tmp_PBR3=[]
        
        #         
        print('Creating ADRC instances')
        # create ADRC isntances
        self.ADRC_PBR1 = ADRC()
        self.ADRC_PBR2 = ADRC()
        self.ADRC_PBR3 = ADRC()
        
        print('Creating Microalgae Models')
        self.MicroalgaePBR1 = Model_Microalgae(D=0, X=0.2, Q=8, S=100)
        self.MicroalgaePBR2 = Model_Microalgae(D=0, X=0.2, Q=8, S=100)
        self.MicroalgaePBR3 = Model_Microalgae(D=0, X=0.2, Q=8, S=100)
            
                
        # Create  Excel Boot to save data
        self.Excel_Date = datetime.datetime.now()
        self.today = self.Excel_Date.strftime("%h.%d.%Y") ############
        self.Hours = self.Excel_Date.strftime("%H.%M.%S") ############
        
        # Define the time after which you want to add a sheet
        self.add_sheet_time = datetime.datetime.strptime("00:01:00", "%H:%M:%S").time()

        ### Create Light Excel Workbook
        workbook = xlsxwriter.Workbook('Light_'+ self.Hours +"_"+ self.today +'.xlsx')
        bold = workbook.add_format({'bold': True})
        worksheet = workbook.add_worksheet()
        worksheet.write('A1', 'Time', bold)
        worksheet.write('B1', 'Light PBR1', bold)
        worksheet.write('C1', 'Light PBR2', bold)
        worksheet.write('D1', 'Light PBR3', bold)
        # save changes to workbook
        workbook.close()
        
        ### Create Nivel Excel Workbook
        workbook = xlsxwriter.Workbook('Nivel_'+ self.Hours +"_"+ self.today +'.xlsx')
        bold = workbook.add_format({'bold': True})
        worksheet = workbook.add_worksheet()
        worksheet.write('A1', 'Time', bold)
        worksheet.write('B1', 'Nivel PBR1', bold)
        worksheet.write('C1', 'Nivel PBR2', bold)
        worksheet.write('D1', 'Nivel PBR3', bold)
        # save changes to workbook
        workbook.close()
        
        ### Create Temperature Excel Workbook
        workbook = xlsxwriter.Workbook('Temperature_'+ self.Hours +"_"+ self.today +'.xlsx')
        bold = workbook.add_format({'bold': True})
        worksheet = workbook.add_worksheet()
        worksheet.write('A1', 'Time', bold)
        worksheet.write('B1', 'Temperature PBR1', bold)
        worksheet.write('C1', 'Temperature PBR2', bold)
        worksheet.write('D1', 'Temperature PBR3', bold)
        workbook.close()
        # save changes to workbook
        
        ### Create I2C Excel Workbook
        workbook = xlsxwriter.Workbook('I2C_'+ self.Hours +"_"+ self.today +'.xlsx')
        bold = workbook.add_format({'bold': True})
        worksheet = workbook.add_worksheet()
        worksheet.write('A1', 'Time', bold)
        worksheet.write('B1', 'PH PBR1', bold)
        worksheet.write('C1', 'pH PBR2', bold)
        worksheet.write('D1', 'pH PBR3', bold)
        worksheet.write('E1', 'DO PBR1', bold)
        worksheet.write('F1', 'DO PBR2', bold)
        worksheet.write('G1', 'DO PBR3', bold)
        # save changes to workbook
        workbook.close()
        
        
        # Variables for start saving excel data
        self.start_row_Ligt_Control = 1
        self.start_row_nivel_monitoring = 1
        self.start_row_IC2_monitoring = 1
        self.start_row_Temperature_Control = 1
        
        # Setpoint
        self.setpoint = 0
        # Initial values of the plant response Inocolum
         
        self.y_PBR1 = 0.2
        self.y_PBR2 = 0.2
        self.y_PBR3 = 0.2
        
        
        
        # Create an entry widget for the setpoint
        self.entry_setpoint = ttk.Entry(self.root)
        self.entry_setpoint.place(x=80, y=435)

       
        # Here is the start buttons
        
        tk.Button(text="Start",command=self.star_FBR,width=30).place(x=80,y=435+150)
        tk.Button(text="Stop",command=self.stop_FBR,width=30).place(x=80,y=465+150)
        
        # Initialize filter coefficients for level monitoring
        self.filter_order = 4
        self.filter_cutoff = 0.05  # Adjust this as needed
        self.filter_b, self.filter_a = butter(self.filter_order, self.filter_cutoff, 'lowpass')
        
        
        # FIlter for light

        #  Set the velocity e PMP
        self.level_Peristaltic_PBR1=False
        self.level_Peristaltic_PBR2=False
        self.level_Peristaltic_PBR3=False
        
        #setpoint    
                
        # Create a label object
        tk.Label( text="Setpoint",font=("Helvetica",15), fg='white',bg=('#00205B')).place(x=10,y=400)
        tk.Label( text="Ref:",font=("Helvetica",12), fg='white', bg=('#00205B')).place(x=10,y=430)
        
        
        # Create a submit button
        button_submit = ttk.Button(text="Submit", command=self.submit_setpoint)
        button_submit.place(x=80, y=485)

        # Create a label to display the setpoint
        self.label_display_setpoint = ttk.Label(self.root, text="Current setpoint: NA")
        self.label_display_setpoint.place(x=80, y=525)
        
        print('Ready to start')
        
    def submit_setpoint(self):
        self.setpoint = self.entry_setpoint.get()
        try:
            self.setpoint = float(self.setpoint)
            self.label_display_setpoint.config(text=f"Current setpoint: {self.setpoint}" + ' (10^9 cells)' )
            # Use the setpoint value in your process
            print(f"Current  setpoint value: {self.setpoint}")
        except ValueError:
            self.label_display_setpoint.config(text="Invalid setpoint value. Please enter a number.")
    
    def MyThread(self, a):
        # Obtén la hora actual
        current_time = datetime.datetime.now()

        # Define las horas a las que se ejecutan las tareas
        task1_hour = 18  # Ejecuta la tarea 1 a las 8:00
        task2_hour = 8  # Ejecuta la tarea 2 a las 0:00 (medianoche)

        # Calcula la hora actual y redondea hacia abajo al múltiplo de 16 más cercano
        current_hour = current_time.hour
        

        # Verifica si es hora de ejecutar la tarea 1
        if current_hour >= task1_hour:
            print("Encentiendo Luces")
            #PBR1
            Labjack1.sendValue('EIO0', True)  
            #PBR2
            Labjack1.sendValue('EIO4', True)  
            #PBR3
            Labjack1.sendValue('EIO6', True)
            # Reset the controllers
              
            self.PID_Light_PBR1.reset()
            self.PID_Light_PBR2.reset()
            self.PID_Light_PBR3.reset()
            
            # Realiza las acciones necesarias para la tarea 1 aquí

        # Verifica si es hora de ejecutar la tarea 2
        elif current_hour >= task2_hour:
            print("Apagando Luces")
            #PBR1
            Labjack1.sendValue('EIO0', False)  # 1
            #PBR2
            Labjack1.sendValue('EIO4', False)  # 2
            #PBR3
            Labjack1.sendValue('EIO6', False)  # 1
            # Realiza las acciones necesarias para la tarea 2 aquí
        
        time.sleep(10)



        
        

        
        
    
    def ThreadADRC(self, dt):
        # print("No hago nada :O")
        # Global self.setpoint
        # Call the ADRC controller instances for PBR1, PBR2, and PBR3
        
        Dilution_PBR1 = self.ADRC_PBR1.ComputeADRC(setpoint=self.setpoint, y=self.y_PBR1, K=15, dt=dt)
        Dilution_PBR2 = self.ADRC_PBR2.ComputeADRC(setpoint=self.setpoint, y=self.y_PBR2, K=15, dt=dt)
        Dilution_PBR3 = self.ADRC_PBR3.ComputeADRC(setpoint=self.setpoint, y=self.y_PBR3, K=15, dt=dt)

        
        # Update the Microalgae Models for PBR1, PBR2, and PBR3 with control_output
        self.y_PBR1, newQ_Pbr1, newS_Pbr1 = self.MicroalgaePBR1.update_tertiolecta(Dilution_PBR1, dt=dt)
        self.y_PBR2, newQ_Pbr2, newS_Pbr2 = self.MicroalgaePBR2.update_tertiolecta(Dilution_PBR2, dt=dt)
        self.y_PBR3, newQ_Pbr3, newS_Pbr3 = self.MicroalgaePBR3.update_tertiolecta(Dilution_PBR3, dt=dt)
        
        #print(self.y_PBR1)
        #print(self.y_PBR2)
        #print(self.y_PBR3)
        #
        #print("Dilution rates")
        #print(D_PBR1)
        #print(D_PBR2)
        #print(D_PBR3)
        
         
# General funtions
    
    def star_FBR(self): 
        print('Creating ADRC instances')
        # create ADRC isntances
        
        print('Starting or Resuming Threads')
        N_samples = 100 # Number of elements to be stored
        
        # Threads FBR
        dt_ADRC = 0.01
        self.thADRC = continuous_threading.PeriodicThread(dt_ADRC, target=self.ThreadADRC, args=[dt_ADRC])
        self.th1 = continuous_threading.ContinuousThread(target=self.MyThread, args=[0] ) #Defining the thread as continuos thread in a loop
        self.th2 = continuous_threading.ContinuousThread(target=self.I2C_monitoring) #Defining the thread as continuos thread in a loop
        self.th3 = continuous_threading.PeriodicThread(self.Nivel_executiontime, target=self.nivel_monitoring, args=[N_samples]) #Defining the thread as periodic thread in a loop
        self.th4 = continuous_threading.PeriodicThread(self.Light_executiontime, target=self.Light_monitoring, args=[N_samples]) #Defining the thread as periodic thread in a loop
        self.th5 = continuous_threading.PeriodicThread(self.pid_temperature_executiontime, target=self.temperature_control, args=[N_samples]) #Defining the thread as periodic thread in a loop
        self.th6 = continuous_threading.PeriodicThread(self.Light_Control_executiontime, target=self.Light_Control, args=[N_samples]) 
        
        # self.th.daemon = True # Set thread to daemon
        #print(self.th.is_running)
        #if self.th.is_running == False:
        self.thADRC.start()
        self.th1.start()
        self.th2.start()
        self.th3.start()
        self.th4.start()
        self.th5.start()
        self.th6.start()
            
    def stop_FBR(self):
        print('Pausing Threads')
        self.thADRC.stop()
        print('Pausing ADRC Thread')
        self.th1.stop()
        print('Pausing My Thread')
        self.th2.stop()
        print('Pausing IC2 Network')
        self.th3.stop()
        print('Pausing Level Monitoring')
        self.th4.stop()
        print('Pausing Light Monitoring')
        self.th5.stop()
        print('Pausing temperature control')
        self.th6.stop()
        print('Pausing Light control')
       
    def main(self):
        # call the main function of each PBR
        self.wFBR1.__main__(self.tab_FBR1)
        self.wFBR2.__main__(self.tab_FBR2)
        self.wFBR3.__main__(self.tab_FBR3)
        
    def update_excel_file_Light(self, sheet, Time, Intensity_PBR1, Intensity_PBR2, Intensity_PBR3):
        
        # Load the Excel workbook
        workbook = openpyxl.load_workbook('Light_'+ self.Hours +"_"+ self.today +'.xlsx')
        # Select the active worksheet
        worksheet = workbook.active
        # Update a cell value
        worksheet["A" + str(sheet)] = Time
        worksheet["B" + str(sheet)] = str(Intensity_PBR1)
        worksheet["C" + str(sheet)] = str(Intensity_PBR2)
        worksheet["D" + str(sheet)] = str(Intensity_PBR3)
        
        # Save the workbook
        workbook.save('Light_'+ self.Hours +"_"+ self.today +'.xlsx')
    
    def update_excel_file_nivel(self, sheet, Time, Nivel_PBR1, Nivel_PBR2, Nivel_PBR3):
        
        # Load the Excel workbook
        workbook = openpyxl.load_workbook('Nivel_'+ self.Hours +"_"+ self.today +'.xlsx')
        # Select the active worksheet
        worksheet = workbook.active
        # Update a cell value
        worksheet["A" + str(sheet)] = Time
        worksheet["B" + str(sheet)] = str(Nivel_PBR1)
        worksheet["C" + str(sheet)] = str(Nivel_PBR2)
        worksheet["D" + str(sheet)] = str(Nivel_PBR3)
        
        # Save the workbook
        workbook.save('Nivel_'+ self.Hours +"_"+ self.today +'.xlsx')
             
    def update_excel_file_temperature(self, sheet, Time, Temperature_PBR1, Temperature_PBR2, Temperature_PBR3):
        
        # Load the Excel workbook
        workbook = openpyxl.load_workbook('Temperature_'+ self.Hours +"_"+ self.today +'.xlsx')
        # Select the active worksheet
        worksheet = workbook.active
        # Update a cell value
        worksheet["A" + str(sheet)] = Time
        worksheet["B" + str(sheet)] = str(Temperature_PBR1)
        worksheet["C" + str(sheet)] = str(Temperature_PBR2)
        worksheet["D" + str(sheet)] = str(Temperature_PBR3)
        
        # Save the workbook
        workbook.save('Temperature_'+ self.Hours +"_"+ self.today +'.xlsx')
        
    def update_excel_file_I2C(self, sheet, Time, ph_PBR1, ph_PBR2, ph_PBR3, DO_PBR1, DO_PBR2, DO_PBR3):
        
        # Load the Excel workbook
        workbook = openpyxl.load_workbook('I2C_'+ self.Hours +"_"+ self.today +'.xlsx')
        # Select the active worksheet
        worksheet = workbook.active
        # Update a cell value
        worksheet["A" + str(sheet)] = Time
        worksheet["B" + str(sheet)] = str(ph_PBR1)
        worksheet["C" + str(sheet)] = str(ph_PBR2)
        worksheet["D" + str(sheet)] = str(ph_PBR3)
        worksheet["E" + str(sheet)] = str(DO_PBR1)
        worksheet["F" + str(sheet)] = str(DO_PBR2)
        worksheet["G" + str(sheet)] = str(DO_PBR3)
        
        # Save the workbook
        workbook.save('I2C_'+ self.Hours +"_"+ self.today +'.xlsx')
        
    def apply_butterworth_filter(self, data):
       return lfilter(self.filter_b, self.filter_a, data) 
    
    def moving_average_filter(self, data, window_size):
        cumsum = np.cumsum(np.insert(data, 0, 0))
        moving_avg = (cumsum[window_size:] - cumsum[:-window_size]) / window_size
        return moving_avg
      
    def get_average_of_last_third(self, data, num_samples):
        last_third = data[-(num_samples // 3):]
        average_last_third = sum(last_third) / len(last_third)
        return average_last_third 
    
    def get_average(self, data):
        if len(data) == 0:
            return 0
        else:
            return sum(data) / len(data)
    
    def PID_tmp_PBR1reset(self):
        self.PID_temp_PBR1.reset()

    def PID_tmp_PBR2reset(self):
        self.PID_temp_PBR2.reset()

    def PID_tmp_PBR3reset(self):
        self.PID_temp_PBR3.reset()
   
# Control funtions

    def nivel_monitoring(self, N):
        # Nivel PBR1
        voltaje_Nivel_PBR1 = float(Labjack2.readValue('AIN0'))
        nivel_real_PBR1 = float(5.8953 * voltaje_Nivel_PBR1 + 2.5354)
        self.wFBR1.xLvl.clear()
        self.wFBR1.xLvl.set_xlabel('$Time$'), self.wFBR1.xLvl.set_ylabel('$cm$')
        self.lvl_PBR1.append(nivel_real_PBR1)
        self.Timelvl_PBR1.append(datetime.datetime.now())
        self.lvl_PBR1=self.lvl_PBR1[-N:]        
        self.Timelvl_PBR1=self.Timelvl_PBR1[-N:]
        
        # Apply Butterworth filter
        filtered_lvl_PBR1 = self.apply_butterworth_filter(self.lvl_PBR1)
        
        # Update plot with filtered data
        self.wFBR1.xLvl.plot(self.Timelvl_PBR1, filtered_lvl_PBR1), self.wFBR1.xLvl.grid(True)
        self.wFBR1.lineLvl.draw()
        
        
        
        # Nivel PBR2
        voltaje_Nivel_PBR2 = float(Labjack2.readValue('AIN1'))
        nivel_real_PBR2 = float(5.8953 * voltaje_Nivel_PBR2 + 2.5354)
        self.wFBR2.xLvl.clear()
        self.wFBR2.xLvl.set_xlabel('$Time$'), self.wFBR2.xLvl.set_ylabel('$cm$')
        self.lvl_PBR2.append(nivel_real_PBR2)
        self.Timelvl_PBR2.append(datetime.datetime.now())
        self.lvl_PBR2=self.lvl_PBR2[-N:]   
        self.Timelvl_PBR2=self.Timelvl_PBR2[-N:]  
        
        # Apply Butterworth filter
        filtered_lvl_PBR2 = self.apply_butterworth_filter(self.lvl_PBR2)
        
        self.wFBR2.xLvl.plot(self.Timelvl_PBR2,filtered_lvl_PBR2), self.wFBR2.xLvl.grid(True)   
        self.wFBR2.lineLvl.draw()  
        
        # Nivel PBR3
        voltaje_Nivel_PBR3 = float(Labjack2.readValue('AIN2'))
        nivel_real_PBR3 = float(5.8953 * voltaje_Nivel_PBR3 + 2.5354)
        self.wFBR3.xLvl.clear()
        self.wFBR3.xLvl.set_xlabel('$Time$'), self.wFBR3.xLvl.set_ylabel('$cm$')
        self.lvl_PBR3.append(nivel_real_PBR3)
        self.Timelvl_PBR3.append(datetime.datetime.now())
        self.lvl_PBR3=self.lvl_PBR3[-N:]    
        self.Timelvl_PBR3=self.Timelvl_PBR3[-N:]  
        
        # Apply Butterworth filter
        filtered_lvl_PBR3 = self.apply_butterworth_filter(self.lvl_PBR3)
        
        self.wFBR3.xLvl.plot(self.Timelvl_PBR3,filtered_lvl_PBR3), self.wFBR3.xLvl.grid(True)  
        self.wFBR3.lineLvl.draw()  
        
        # Control_Nivel on/off simply
        # First computed the average of the last elements of the vectors
        last_third_PBR1 = filtered_lvl_PBR1[-(N//3):]
        average_last_third_PBR1 = sum(last_third_PBR1) / len(last_third_PBR1)
        
        last_third_PBR2 = filtered_lvl_PBR2[-(N//3):]
        average_last_third_PBR2 = sum(last_third_PBR2) / len(last_third_PBR2)
        
        last_third_PBR3 = filtered_lvl_PBR3[-(N//3):]
        average_last_third_PBR3 = sum(last_third_PBR3) / len(last_third_PBR3)
        
        # Control conditions on or off peristaltic PMPS
        if (average_last_third_PBR1 > self.wFBR1.Ref_lvl.get()):
            self.level_Peristaltic_PBR1=True
        else: 
            self.level_Peristaltic_PBR1=False
            
        if (average_last_third_PBR2 > self.wFBR2.Ref_lvl.get()):
            self.level_Peristaltic_PBR2=True
        else: 
            self.level_Peristaltic_PBR2=False
            
        if (average_last_third_PBR3 > self.wFBR3.Ref_lvl.get()):
            self.level_Peristaltic_PBR3=True
        else: 
            self.level_Peristaltic_PBR3=False
        

        # Update excel File
        Time = datetime.datetime.now()
        Time = time.strftime("%H.%M.%S")
        self.start_row_nivel_monitoring += 1
        self.update_excel_file_nivel(sheet=self.start_row_nivel_monitoring,Time=Time, Nivel_PBR1=nivel_real_PBR1, Nivel_PBR2=nivel_real_PBR2, Nivel_PBR3=nivel_real_PBR3 )

    def temperature_control(self, N):
        #z is the time of execution of the thread
        

        # Reset the controllers if the setpoint change in menus

        
        # append time temperature 
        self.TimeTemperature.append(datetime.datetime.now())
        self.TimeTemperature = self.TimeTemperature[-N:]
        
        # https://labjack.com/pages/support?doc=%2Fdatasheets%2Faccessories%2Fei-1034-datasheet%2F
        # °C = (55.56*volts) + 255.37 - 273.15
        
        # Temperature PBR1
        Temp_PBR1 = 55.56*Labjack1.readValue('AIN0') + 255.37 - 273.15 
        self.wFBR1.xTemp.clear()
        self.wFBR1.xTemp.set_xlabel('$Time$'), self.wFBR1.xTemp.set_ylabel('$°C$')
        self.Tmp_PBR1.append(Temp_PBR1)
        self.Tmp_PBR1=self.Tmp_PBR1[-N:]        
        # Filter
        filtered_temp_PBR1 = self.apply_butterworth_filter(self.Tmp_PBR1)
        self.wFBR1.xTemp.plot(self.TimeTemperature, filtered_temp_PBR1), self.wFBR1.xTemp.grid(True)
        self.wFBR1.lineTemp.draw()
        
        # Control temp PBR1
        
        # Setpoint
        self.PID_temp_PBR1.setPoint(self.wFBR1.Ref_tmp.get())
        # Call update function of the PID and send the value of the actual temperature
        # Control_Nivel on/off simply
        # First computed the average of the last elements of the vectors
        average_last_third_PBR1 = self.get_average_of_last_third(filtered_temp_PBR1, N)
        UPID_Temp_PBR1 = self.PID_temp_PBR1.update(average_last_third_PBR1)
        
        # Write the temperature computed value in the labjack
        Labjack1.sendValue('DAC0', np.interp(UPID_Temp_PBR1, [0, 100], [0, 5])) # Interp
        
        # Turn on Cooler
        if Temp_PBR1 > self.wFBR1.Ref_tmp.get(): #VENTILADOR ACTIVACIÓN DIGITAL
                Labjack1.sendValue('CIO1',5)
        else:
                Labjack1.sendValue('CIO1',0)
        
                
                
        # Temperature PBR2
        Temp_PBR2 = 55.56*Labjack1.readValue('AIN7') + 255.37 - 273.15 
        self.wFBR2.xTemp.clear()
        self.wFBR2.xTemp.set_xlabel('$Time$'), self.wFBR2.xTemp.set_ylabel('$°C$')
        self.Tmp_PBR2.append(Temp_PBR2)
        self.Tmp_PBR2=self.Tmp_PBR2[-N:]
        filtered_temp_PBR2 = self.apply_butterworth_filter(self.Tmp_PBR2)        
        self.wFBR2.xTemp.plot(self.TimeTemperature, filtered_temp_PBR2), self.wFBR2.xTemp.grid(True)
        self.wFBR2.lineTemp.draw()   
        
        # Control temp PBR2
        
        # Setpoint
        self.PID_temp_PBR2.setPoint(self.wFBR2.Ref_tmp.get())
        # Call update function of the PID and send the value of the actual temperature
        
        average_last_third_PBR2 = self.get_average_of_last_third(filtered_temp_PBR2, N)
        UPID_Temp_PBR2 = self.PID_temp_PBR2.update(average_last_third_PBR2)
           
        # Write the temperature computed value in the labjack
        #Labjack1.sendValue('DAC1', np.interp(UPID_Temp_PBR2, [0, 100], [0, 5])) # Interp
        
        # Turn on Cooler
        if Temp_PBR2 > self.wFBR2.Ref_tmp.get(): 
                Labjack1.sendValue('CIO2',5)
        else:
                Labjack1.sendValue('CIO2',0)
        
        
        # Temperature PBR3
        Temp_PBR3 = 55.56*Labjack1.readValue('AIN3') + 255.37 - 273.15 
        self.wFBR3.xTemp.clear()
        self.wFBR3.xTemp.set_xlabel('$Time$'), self.wFBR3.xTemp.set_ylabel('$°C$')
        self.Tmp_PBR3.append(Temp_PBR3)
        self.Tmp_PBR3=self.Tmp_PBR3[-N:]     
        filtered_temp_PBR3 = self.apply_butterworth_filter(self.Tmp_PBR3)     
        self.wFBR3.xTemp.plot(self.TimeTemperature, filtered_temp_PBR3), self.wFBR3.xTemp.grid(True)
        self.wFBR3.lineTemp.draw()

        # Control temp PBR3
        
        # Setpoint
        self.PID_temp_PBR3.setPoint(self.wFBR3.Ref_tmp.get())
        # Call update function of the PID and send the value of the actual temperature
        average_last_third_PBR3 = self.get_average_of_last_third(filtered_temp_PBR3, N)
        #print("Medido PBR3", average_last_third_PBR3)
        UPID_Temp_PBR3 = self.PID_temp_PBR3.update(average_last_third_PBR3)

        #Labjack1.sendValue('TDAC2', np.interp(UPID_Temp_PBR3, [0, 100], [0, 5])) # Interp
        
        # Turn on Cooler
        if Temp_PBR3 > self.wFBR3.Ref_tmp.get(): #VENTILADOR ACTIVACIÓN DIGITAL
                Labjack1.sendValue('CIO3',5)
        else:
                Labjack1.sendValue('CIO3',0)
        
          
        # Update excel File
        Time = datetime.datetime.now()
        Time = time.strftime("%H.%M.%S")
        self.start_row_Temperature_Control += 1
        self.update_excel_file_temperature(sheet=self.start_row_Temperature_Control,Time=Time, Temperature_PBR1=Temp_PBR1, Temperature_PBR2=Temp_PBR2, Temperature_PBR3=Temp_PBR3)
    
    def Light_Control(self, N):
        
        average_last_third_PBR1 = self.get_average(self.avfiltered_Iin_PBR1)
        average_last_third_PBR2 = self.get_average(self.avfiltered_Iin_PBR2)
        average_last_third_PBR3 = self.get_average(self.avfiltered_Iin_PBR3)
        
        
        
        
        # Light Control PBR1
        # Setpoint
        self.PID_Light_PBR1.setPoint(self.wFBR1.Ref_luz.get())
        # Call update function of the PID 
        PID_Light_PBR1 = self.PID_Light_PBR1.update(average_last_third_PBR1)
        Labjack1.sendValue('TDAC3', np.interp(PID_Light_PBR1, [0, 1000], [3, 5])) # Interp
        
        # Light Control PBR2
        # Setpoint
        self.PID_Light_PBR2.setPoint(self.wFBR2.Ref_luz.get())
        # Call update function of the PID 
        PID_Light_PBR2 = self.PID_Light_PBR2.update(average_last_third_PBR2)
        Labjack1.sendValue('TDAC4', np.interp(PID_Light_PBR2, [0, 1000], [3, 5])) # Interp
        
        # Light Control PBR3
        # Setpoint
        self.PID_Light_PBR3.setPoint(self.wFBR3.Ref_luz.get())
        PID_Light_PBR3 = self.PID_Light_PBR3.update(current_value=average_last_third_PBR3)
        Labjack1.sendValue('TDAC5', np.interp(PID_Light_PBR3, [0, 1000], [3, 5])) # Interp
        
        
    
    def Light_monitoring(self, N):
        
        # Append time vector
        self.TimeIin.append(datetime.datetime.now())
        self.TimeIin=self.TimeIin[-N:]
        
        
        # For more information about this equation refer to manual of the Sensor S2-141, taking into account that  you have to multiply your readings by 1.25 to account for the immersion effect due to the light refracting when it hits the water
        
        # PBR1
        self.Intensity_PBR1 = ((Labjack1.readValue('AIN5')-0.399)*100000)*1.25  #Read analoge input
        self.wFBR1.xIin.clear()
        self.wFBR1.xIin.grid(True),self.wFBR1.xIin.set_xlabel('$Time$'),self.wFBR1.xIin.set_ylabel('$\mu mol \cdot m^{-2} \cdot s^{-1}$')
        # Append Light vector
        self.Iin_PBR1.append(self.Intensity_PBR1)
        # Limits vector to have 20 elements
        self.Iin_PBR1=self.Iin_PBR1[-N:] 
        self.filtered_Iin_PBR1 = self.apply_butterworth_filter(self.Iin_PBR1)

        # PBR2
        self.Intensity_PBR2 = ((Labjack1.readValue('AIN13')-0.399)*100000)*1.25  #Read analoge input
        self.wFBR2.xIin.clear()
        self.wFBR2.xIin.grid(True),self.wFBR2.xIin.set_xlabel('$Time$'),self.wFBR2.xIin.set_ylabel('$\mu mol \cdot m^{-2} \cdot s^{-1}$')
        # Append Light vector
        self.Iin_PBR2.append(self.Intensity_PBR2)
        # Limits vector to have 20 elements
        self.Iin_PBR2=self.Iin_PBR2[-N:] 
        self.filtered_Iin_PBR2 = self.apply_butterworth_filter(self.Iin_PBR2)
        
        
        
        # PBR3
        self.Intensity_PBR3 = ((Labjack1.readValue('AIN9')-0.399)*100000)*1.25 #Read analoge input
        self.wFBR3.xIin.clear()
        self.wFBR3.xIin.grid(True),self.wFBR3.xIin.set_xlabel('$Time$'),self.wFBR3.xIin.set_ylabel('$\mu mol \cdot m^{-2} \cdot s^{-1}$')
        # Append Light vector
        self.Iin_PBR3.append(self.Intensity_PBR3)
        # Limits vector to have 20 elements
        self.Iin_PBR3=self.Iin_PBR3[-N:] 
        self.filtered_Iin_PBR3 = self.apply_butterworth_filter(self.Iin_PBR3)       
        
 
        
        # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #
        self.avfiltered_Iin_PBR1 =self.moving_average_filter(self.Iin_PBR1, 3)
        self.avfiltered_Iin_PBR2 =self.moving_average_filter(self.Iin_PBR2, 3)
        self.avfiltered_Iin_PBR3 =self.moving_average_filter(self.Iin_PBR3, 3)
        
        self.average_Iin_PBR1.append(self.get_average(self.avfiltered_Iin_PBR1)) 
        self.average_Iin_PBR2.append(self.get_average(self.avfiltered_Iin_PBR2)) 
        self.average_Iin_PBR3.append(self.get_average(self.avfiltered_Iin_PBR3)) 
        
        self.average_Iin_PBR1 = self.average_Iin_PBR1[-N:]
        self.average_Iin_PBR2 = self.average_Iin_PBR2[-N:]
        self.average_Iin_PBR3 = self.average_Iin_PBR3[-N:]
        
        #####################################################################
        
        #### plot graphs
        self.wFBR1.xIin.plot(self.TimeIin, self.average_Iin_PBR1)
        self.wFBR1.xIin.set_ylim([0, 600])
        self.wFBR1.lineIin.draw()
        
        self.wFBR2.xIin.plot(self.TimeIin, self.average_Iin_PBR2)
        self.wFBR2.xIin.set_ylim([0, 600])
        self.wFBR2.lineIin.draw()
        
        self.wFBR3.xIin.set_ylim([0, 600])
        self.wFBR3.xIin.plot(self.TimeIin, self.average_Iin_PBR3)
        self.wFBR3.lineIin.draw()
        
        # Update excel File
        Time = datetime.datetime.now()
        Time = time.strftime("%H.%M.%S")
        self.start_row_Ligt_Control += 1
        self.update_excel_file_Light(sheet=self.start_row_Ligt_Control,Time=Time, Intensity_PBR1=self.Intensity_PBR1, Intensity_PBR2=self.Intensity_PBR2, Intensity_PBR3=self.Intensity_PBR3)
        
        
    
    def I2C_monitoring(self):
   
        delay = 0.2
        Response_time_pH = 0.9
        Reponse_time_DO = 0.6
        Response_time_Ezo = 0.3
        
        # start monitoring 
        
        # DO PBR1
        Labjack1.initI2C(1, 0, 6) #EL OBJETO comunicación LLAMA AL MÉTODO initI2C (TX,RX,DIRECCIÓN)
        Labjack1.sendValueI2C([82], delay=Reponse_time_DO) #EL OBJETO comunicación LLAMA AL MÉTODO sendValueI2C ([COMANDO ASCII]) EN CASO DE NECESITAR OTRA UTILIDAD VER MANUAL DEL SENSOR
        time.sleep(delay) # Aditional required delay
        
        # This is in case of a error in the comunication
        try:
            DO_real_PBR1=float(Labjack1.readValueI2C()) #L read the sensor
        except:
            DO_real_PBR1=self.DO_PBR1.pop() #las value
            print("Error de lectura DO PBR1")
        else:     # The else block lets you execute code when there is no error.
            pass
        finally:  # Block lets you execute code, regardless of the result of the try- and except block
            pass
        
        
        
        self.wFBR1.xDO.clear()
        self.wFBR1.xDO.set_xlabel('$Time$'),self.wFBR1.xDO.set_ylabel('$mg \cdot L^{-1}$')
        
        # Comparing len        
        if len(self.Time_PBR1_DO) == len(self.DO_PBR1):
            self.wFBR1.xDO.plot(self.Time_PBR1_DO,self.DO_PBR1), self.wFBR1.xDO.grid(True)
        elif len(self.Time_PBR1_DO) > len(self.DO_PBR1):
            self.Time_PBR1_DO.popleft()  # remove the first value if the list is full
        else:
            self.DO_PBR1.popleft()
        # apending data
        self.Time_PBR1_DO.append(datetime.datetime.now()), self.DO_PBR1.append(DO_real_PBR1)
        self.wFBR1.xDO.set_ylim(0, 50)
        self.wFBR1.lineDO.draw()
    
                
        # pH PBR1
        Labjack1.initI2C(1, 0, 3)
        Labjack1.sendValueI2C([82], delay=Response_time_pH)#114
        
        time.sleep(delay) #POR NADA DEL MUNDO SE PUEDE CAMBIAR ESTE RETARDO (VER MANUAL DEL SENSOR)
        
        # This is in case of a error in the comunication
        try:
            pH_real_PBR1=float(Labjack1.readValueI2C())
        except:
            pH_real_PBR1=self.pH_PBR1.pop() # Last value
            print("Error de lectura PH PBR1")
                
        #print('pH1',pH_real_PBR1)
        self.wFBR1.xpH.clear()
        self.wFBR1.xpH.set_xlabel('$Time$'),self.wFBR1.xpH.set_ylabel('$pH$')
        
        # Comparing len        
        if len(self.Time_PBR1_pH) == len(self.pH_PBR1):
            self.wFBR1.xpH.plot(self.Time_PBR1_pH,self.pH_PBR1), self.wFBR1.xpH.grid(True)
        elif len(self.Time_PBR1_pH) > len(self.pH_PBR1):
            self.Time_PBR1_pH.popleft()  # remove the first value if the list is full
        else:
            self.pH_PBR1.popleft()
        # apending data
        self.Time_PBR1_DO.append(datetime.datetime.now()), self.DO_PBR1.append(DO_real_PBR1)
        self.wFBR1.xDO.set_ylim(0, 50)
        self.wFBR1.lineDO.draw()
    
        
        
        self.Time_PBR1_pH.append(datetime.datetime.now()), self.pH_PBR1.append(pH_real_PBR1) 
        self.wFBR1.xpH.set_ylim(0, 14)
        self.wFBR1.linepH.draw()

        # DO PBR2
        Labjack1.initI2C(1, 0, 5)
        Labjack1.sendValueI2C([82], delay=Reponse_time_DO)
        time.sleep(delay)
        # This is in case of a error in the comunication
        try:
            DO_real_PBR2=float(Labjack1.readValueI2C())
        except:
            DO_real_PBR2=self.DO_PBR2.pop() # Last value 
            print("Error de lectura DO PBR2")
        
        self.wFBR2.xDO.clear()
        self.wFBR2.xDO.set_xlabel('$Time$'),self.wFBR2.xDO.set_ylabel('$mg/L$')
        
        # Comparing len        
        if len(self.Time_PBR2_DO) == len(self.DO_PBR2):
            self.wFBR2.xDO.plot(self.Time_PBR2_DO,self.DO_PBR2), self.wFBR2.xDO.grid(True)
        elif len(self.Time_PBR2_DO) > len(self.DO_PBR2):
            self.Time_PBR2_DO.popleft()  # remove the first value if the list is full
        else:
            self.DO_PBR2.popleft()
        # apending data
        
        
        self.Time_PBR2_DO.append(datetime.datetime.now()), self.DO_PBR2.append(DO_real_PBR2)  
        self.wFBR2.xDO.set_ylim(0, 50)
        self.wFBR2.lineDO.draw()
  
  
        
        # pH PBR2
        Labjack1.initI2C(1, 0, 2)
        Labjack1.sendValueI2C([82], delay=Response_time_pH) # Command to read
        time.sleep(delay)
        
        # This is in case of a error in the comunication
        try:
            pH_real_PBR2=float(Labjack1.readValueI2C())
        except:
            pH_real_PBR2=self.pH_PBR2.pop() # last value
            print("Error de lectura pH PBR2")
        
        #print('pH2', pH_real_PBR2)
        self.wFBR2.xpH.clear()
        self.wFBR2.xpH.set_xlabel('$Time$'),self.wFBR2.xpH.set_ylabel('$pH$')
        # Comparing len        
        if len(self.Time_PBR2_pH) == len(self.pH_PBR2):
            self.wFBR2.xpH.plot(self.Time_PBR2_pH,self.pH_PBR2), self.wFBR2.xpH.grid(True)
        elif len(self.Time_PBR2_pH) > len(self.pH_PBR2):
            self.Time_PBR2_pH.popleft()  # remove the first value if the list is full
        else:
            self.pH_PBR2.popleft()
        # apending data
        self.Time_PBR2_pH.append(datetime.datetime.now()), self.pH_PBR2.append(pH_real_PBR2)
        self.wFBR2.xpH.set_ylim(0, 14)
        self.wFBR2.linepH.draw()
        
    
        # DO PBR3
        Labjack1.initI2C( 1, 0, 4)
        Labjack1.sendValueI2C([82], delay=Reponse_time_DO)
        time.sleep(delay)
        
        # This is in case of a error in the comunication
        try:
            DO_real_PBR3=float(Labjack1.readValueI2C())
        except:
            DO_real_PBR3=self.DO_PBR3.pop() # last value
            print("Error de lectura DO PBR3")
        
        #print('DO3',DO_real_PBR3)
        self.wFBR3.xDO.clear()
        self.wFBR3.xDO.set_xlabel('$Time$'),self.wFBR3.xDO.set_ylabel('$mg/L$')
         
        # Comparing len        
        if len(self.Time_PBR3_DO) == len(self.DO_PBR3):
            self.wFBR3.xDO.plot(self.Time_PBR3_DO,self.DO_PBR3), self.wFBR3.xDO.grid(True)
        elif len(self.Time_PBR3_DO) > len(self.DO_PBR3):
            self.Time_PBR3_DO.popleft()  # remove the first value if the list is full
        else:
            self.DO_PBR3.popleft()
        # apending data
        
     
        self.Time_PBR3_DO.append(datetime.datetime.now()), self.DO_PBR3.append(DO_real_PBR3)
        self.wFBR3.xDO.set_ylim(0, 50)
        self.wFBR3.lineDO.draw()
        
        # pH PBR3
        Labjack1.initI2C(1, 0, 1)
        Labjack1.sendValueI2C([82], delay=Response_time_pH) #114
        time.sleep(delay)
        # This is in case of a error in the comunication
        try:
            pH_real_PBR3=float(Labjack1.readValueI2C())
        except:
            pH_real_PBR3=self.pH_PBR3.pop() #last value
            print("Error de lectura pH PBR3")
            
            
        self.wFBR3.xpH.clear()
        self.wFBR3.xpH.set_xlabel('$Time$'),self.wFBR3.xpH.set_ylabel('$pH$')
        #print('pH3', pH_real_PBR3)
        # Comparing len        
        if len(self.Time_PBR3_pH) == len(self.pH_PBR3):
            self.wFBR3.xpH.plot(self.Time_PBR3_pH, self.pH_PBR3), self.wFBR3.xpH.grid(True)
        elif len(self.Time_PBR3_pH) > len(self.pH_PBR3):
            self.Time_PBR3_pH.popleft()  # remove the first value if the list is full
        else:
            self.pH_PBR3.popleft()
        # apending data
        
        self.Time_PBR3_pH.append(datetime.datetime.now()), self.pH_PBR3.append(pH_real_PBR3)
        self.wFBR3.xpH.set_ylim(0, 14)
        self.wFBR3.linepH.draw()
        
        ########################################################################
        # Update excel File
        Time = datetime.datetime.now()
        Time = time.strftime("%H.%M.%S")
        self.start_row_IC2_monitoring += 1
        self.update_excel_file_I2C(sheet=self.start_row_IC2_monitoring,Time=Time, ph_PBR1=pH_real_PBR1, ph_PBR2=pH_real_PBR2, ph_PBR3=pH_real_PBR3, DO_PBR1=DO_real_PBR1, DO_PBR2=DO_real_PBR2, DO_PBR3=DO_real_PBR3)
        
        
        ########################################################################
        ########################################################################
        # Update control signals of peristaltic pumps
        ########################################################################
        self.Dilution_rate_PBR1=100
        self.Dilution_rate_PBR2=75
        self.Dilution_rate_PBR3=80
        
        # Definir el valor del controlador, maximal flow rate is 50.05 ml/min 
        D_PBR1 = int(np.interp(self.Dilution_rate_PBR1, [0, 100], [0, 50.05]))
        D_PBR2 = int(np.interp(self.Dilution_rate_PBR2, [0, 100], [0, 50.05]))
        D_PBR3 = int(np.interp(self.Dilution_rate_PBR3, [0, 100], [0, 50.05])) 
        
        #Convertir a litros por dia 72 listros por dia
        #Reactor 3 litros, D es inverso de dias D = 72/3 = 24 
        
        #mu_max = 1.7
        
        
        #This commandas will maintain Constant flow rate DC,[ml/min],[for this much time] 

        # PBR1
        
        # First stop all the pummps 
        
        
        # Crear el comando con el valor del controlador
        
        Labjack1.initI2C(1, 0, 7) 
        Labjack1.sendValueI2C([88], num_bytes_to_read=1, delay=Response_time_Ezo) # X
        Labjack1.sendValueI2C([ord(character) for character in "DC," + str(D_PBR1) +",1"], num_bytes_to_read=1, delay=Response_time_Ezo)
        #print("action PBR1", self.level_Peristaltic_PBR1)
        
        
        # PBR2
        # Crear el comando con el valor del controlador
        
        Labjack1.initI2C(1, 0, 9) 
        Labjack1.sendValueI2C([88], num_bytes_to_read=1, delay=Response_time_Ezo) # X
        Labjack1.sendValueI2C([ord(character) for character in "DC," + str(D_PBR2) +",1"], num_bytes_to_read=1, delay=Response_time_Ezo) #EL OBJETO comunicación LLAMA AL MÉTODO sendValueI2C ([COMANDO ASCII]) EN CASO DE NECESITAR OTRA UTILIDAD VER MANUAL DEL SENSOR
        #print("action PBR2", self.level_Peristaltic_PBR2)
        
        #PBR3        
        # Crear el comando con el valor del controlador
        
        Labjack1.initI2C(1, 0, 11) 
        Labjack1.sendValueI2C([88], num_bytes_to_read=1, delay=Response_time_Ezo) # X
        Labjack1.sendValueI2C([ord(character) for character in "DC," + str(D_PBR3) +",1"], num_bytes_to_read=1, delay=Response_time_Ezo) #EL OBJETO comunicación LLAMA AL MÉTODO sendValueI2C ([COMANDO ASCII]) EN CASO DE NECESITAR OTRA UTILIDAD VER MANUAL DEL SENSOR
        #print("action PBR3", self.level_Peristaltic_PBR3)
        
        # PMP for level
        # Continuous dispensing
        comando_PBR = "D,*"
        
        Labjack1.initI2C(1, 0, 8) 
        if (self.level_Peristaltic_PBR1==True):
            Labjack1.sendValueI2C([ord(character) for character in comando_PBR], num_bytes_to_read=1, delay=Response_time_Ezo)
        if (self.level_Peristaltic_PBR1==False): 
            Labjack1.sendValueI2C([88], num_bytes_to_read=1, delay=Response_time_Ezo) # X  stop command  
                       
        Labjack1.initI2C(1, 0, 10)
        if (self.level_Peristaltic_PBR2==True):
            Labjack1.sendValueI2C([ord(character) for character in comando_PBR], num_bytes_to_read=1, delay=Response_time_Ezo)
        if (self.level_Peristaltic_PBR2==False): 
            Labjack1.sendValueI2C([88], num_bytes_to_read=1, delay=Response_time_Ezo) # X  stop command  
         
        Labjack1.initI2C(1, 0, 12)
        if (self.level_Peristaltic_PBR3==True):
            Labjack1.sendValueI2C([ord(character) for character in comando_PBR], num_bytes_to_read=1, delay=Response_time_Ezo)
        if (self.level_Peristaltic_PBR3==False): 
            Labjack1.sendValueI2C([88], num_bytes_to_read=1, delay=Response_time_Ezo) # X   stop command
            
               
        """Ask for the maximal velocities if required 
        """        
        #comando_PBR = "DC,?"
        #Labjack1.initI2C(1, 0, 7)
        #Labjack1.sendValueI2C([ord(character) for character in comando_PBR], num_bytes_to_read=1, delay=Response_time_Ezo)
        #max = Labjack1.readMaxI2CEZO()
        #print("Max EZO1",max)
        

        #Labjack1.initI2C(1, 0, 9)
        #Labjack1.sendValueI2C([ord(character) for character in comando_PBR], num_bytes_to_read=1, delay=Response_time_Ezo)
        #max = Labjack1.readMaxI2CEZO()
        #print("Max EZO2",max)
        

        #Labjack1.initI2C(1, 0, 11)
        #Labjack1.sendValueI2C([ord(character) for character in comando_PBR], num_bytes_to_read=1, delay=Response_time_Ezo)
        #max = Labjack1.readMaxI2CEZO()
        #print("Max EZO3",max)
   
# Menus information
       
    def f_acerca(self):
        acerca = tk.Toplevel()
        acerca.geometry("700x720")
        acerca.bg=('red')
        acerca.resizable(width=False, height=False)
        
        acerca.title("Acknowledgment:")
        marco1 = ttk.Frame(acerca,padding=(10, 10, 10, 10))
        marco1.pack(side="top", fill="both", expand=True)
        self.fondo=tk.PhotoImage(file="logo.png")
        self.labelfondo=tk.Label(marco1,image=self.fondo).place(x=0,y=30)
        boton1 = tk.Button(marco1, text="Exit",command=acerca.destroy)
        boton1.pack(side="bottom", padx=10, pady=0)
        boton1.focus_set()
        
    def f_about(self):
        about = tk.Toplevel()
        about.geometry("900x720")
        about.bg=('red')
        about.resizable(width=False, height=False)
        about.title("This software was developed by:")
        marco1 = ttk.Frame(about,padding=(10, 10, 10, 10))
        marco1.pack(side="top", fill="both", expand=True)
        self.fondo=tk.PhotoImage(file="logo_II.png")
        self.labelfondo=tk.Label(marco1,image=self.fondo).place(x=0,y=30)
        boton1 = tk.Button(marco1, text="Exit",command=about.destroy)
        boton1.pack(side="bottom", padx=10, pady=0)
        boton1.focus_set()      
  
    def run(self):
        self.root.mainloop()
            
# close the application
    def close(self):
        self.root.destroy()
        # Puertos de colores de las luces Rojo, Verde Azul
        Labjack1.sendValue('EIO0', False)
        Labjack1.sendValue('EIO1', False)
        Labjack1.sendValue('EIO2', False)        
        Labjack1.sendValue('TDAC3',0)
        Labjack1.sendValue('DAC0',0)
        Labjack1.sendValue('CIO1',0)
        Labjack2.sendValue('DAC0',0)
        Labjack2.sendValue('DAC1',0)
        # Puertos de colores de las luces Rojo, Verde Azul
        Labjack1.sendValue('EIO3', False)
        Labjack1.sendValue('EIO4', False)
        Labjack1.sendValue('EIO5', False)
        Labjack1.sendValue('TDAC4', 0)
        Labjack1.sendValue('DAC1', 0)
        Labjack1.sendValue('CIO2', 0)
        Labjack2.sendValue('TDAC0', 0)
        Labjack2.sendValue('TDAC1', 0)
        # Puertos de colores de las luces Rojo, Verde Azul
        Labjack1.sendValue('EIO6', False)
        Labjack1.sendValue('EIO7', False)
        Labjack1.sendValue('CIO0', False)
        Labjack1.sendValue('TDAC5', 0)
        Labjack1.sendValue('TDAC2', 0)
        Labjack1.sendValue('CIO3', 0)
        Labjack2.sendValue('TDAC2', 0)
        Labjack2.sendValue('TDAC3', 0)
        print("Ending program")
        Labjack1.close()
        Labjack2.close()



if __name__ == "__main__":
    main = Main()  # intance of main
    main.main()    # call method main of main class
    main.run()     # call mainloop of the wimdows
    

    


